q='html.parser'
p='Groupe'
o='titre module'
n=enumerate
m=input
f='cadrenormal'
e='table'
d=False
c='D'
b='C'
a='B'
Z='A'
Y='N° module'
X=open
W=len
V=int
U=float
L='Dispositifs'
K='N° dispo.'
J=True
F='Modules'
D=print
C='Dates'
A=''
import re as B
from pathlib import Path as M
import pandas as E
from openpyxl import load_workbook as r
from openpyxl.styles import Border as s,Side,Alignment as t
import time
from datetime import datetime as u
import sys,math,browser_cookie3 as g,requests as h
from bs4 import BeautifulSoup as i
def v(step,total_steps,bar_width=60,title=A,print_perc=J):
	F=title;import sys;G=['█','▏','▎','▍','▌','▋','▊','█'];C=100*U(step)/U(total_steps);H=bar_width*8;E=V(round(C/100*H));J=E/8;I=E%8;B=D=A;D+=G[0]*V(J)
	if I>0:D+=G[I]
	D+='▒'*V(H/8-U(E)/8.)
	if W(F)>0:B=F+': '
	B+='\x1b[0;32m';B+=D;B+='\x1b[0m'
	if print_perc:
		if C>1e2:C=1e2
		B+=' {:6.2f}'.format(C)+' %'
	sys.stdout.write('\r'+B);sys.stdout.flush()
N=['https://gaia.phm.education.gouv.fr/gaia/gacmfgest/animPedagogiques/liste_dispoAnim.jsp','https://gaia.in.phm.education.gouv.fr/gaia/gacmfgest/animPedagogiques/liste_dispoAnim.jsp']
w=['https://gaia.phm.education.gouv.fr/gaia/gacmfgest/dispo/arbre_dispositif.jsp?cCode=','https://gaia.in.phm.education.gouv.fr/gaia/gacmfgest/dispo/arbre_dispositif.jsp?cCode=']
x={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:131.0) Gecko/20100101 Firefox/131.0'}
G=[[K,'Circo','titre dispo.']]
O=[[K,Y,o,"nombre d'heures"]]
I=[[K,Y,p,'date','Heure début','Heure fin']]
AC=[]
P='(E R R E U R|authentification)'
def AD(namefile,data):
	with X(namefile,'w',newline='\r\n',encoding='utf-8')as A:
		for B in data:A.write(f"{B}\n")
def Q(url):
	C=g.firefox()
	try:A=h.get(url,headers=x,cookies=C);A.raise_for_status();B=A.text
	except h.exceptions.RequestException as D:B='exception ERREUR authentification'
	return B
def y():
	A=Q(N[0])
	if B.search(P,A):
		A=Q(N[1])
		if B.search('exception',A):return 0
		else:return 1
	return 0
def z(page):
	A=B.search(P,page)
	if A:D('\nERREUR, ouvrez une session GAIA sous Firefox, vous êtes actuellement non connecté');D("Si l'erreur persiste cliquer dans Stagiaires=>Personnes ou naviguer sur le site GAIA");m('\nAppuyez sur Entrée pour fermer le programme...');sys.exit()
def A0(wb):A=wb[C];A.column_dimensions[Z].width=13;A.column_dimensions[a].width=14;A.column_dimensions[b].width=12;A.column_dimensions[c].width=11;A.column_dimensions['E'].width=12;A.column_dimensions['F'].width=12;A.column_dimensions['G'].width=60;A.column_dimensions['H'].width=14
def A1(wb):A=wb[L];A.column_dimensions[Z].width=13;A.column_dimensions[a].width=14;A.column_dimensions[b].width=80
def A2(wb):A=wb[F];A.column_dimensions[Z].width=13;A.column_dimensions[a].width=10;A.column_dimensions[b].width=70;A.column_dimensions[c].width=16
def R(ws,first_row,first_col,last_col):
	for A in ws.iter_rows(min_row=first_row,max_row=ws.max_row,min_col=first_col,max_col=last_col):
		for B in A:B.alignment=t(horizontal='center')
def S(ws,first_row,first_col):
	A=Side(border_style='thin')
	for B in ws.iter_rows(min_row=first_row,max_row=ws.max_row,min_col=first_col,max_col=ws.max_column):
		for C in B:C.border=s(left=A,right=A,top=A,bottom=A)
def T(ws,ref):A=ws.auto_filter;A.ref=ref
def A3(ws,col):
	for A in ws[col]:
		if A.row>1:A.number_format='DD/MM/YYYY'
def AE(df,num_module):
	B=df.loc[df[Y]==num_module,K]
	if not B.empty:return B.iloc[0]
	return A
def A4(namefile):
	B=namefile;H=I[1:];H.sort(key=lambda x:u.strptime(x[3],'%d/%m/%Y'));D={L:E.DataFrame(G[1:],columns=G[0]),F:E.DataFrame(O[1:],columns=O[0]),C:E.DataFrame(H,columns=I[0])};J=[];K=[]
	for M in range(1,W(I)):J.append(f"=VLOOKUP(B{M+1},Modules!B:D,2,FALSE)");K.append(f"=VLOOKUP(A{M+1},Dispositifs!A:C,2,FALSE)")
	D[C][o]=E.Series(J);D[C]['circo']=E.Series(K)
	with E.ExcelWriter(B,engine='openpyxl')as N:
		for(P,Q)in D.items():Q.to_excel(N,sheet_name=P,index=d)
	A=r(B);A0(A);R(A[C],1,2,6);S(A[C],1,1);A3(A[C],c);T(A[C],'A:H');A1(A);S(A[L],1,1);T(A[L],'A:C');A2(A);R(A[F],1,2,2);R(A[F],1,4,4);S(A[F],1,1);T(A[F],'A:D');A.save(B)
def j(url,from_disk=J):
	E=url;D=A;H=B.search('/([^/]+)\\.jsp\\S*',E);I=B.findall('=([^/]+)',E)
	if H:
		F='cache/'+H.group(1)
		if I:F=F+'_'+I[0]
		D=M(__file__).parent/f"{F}.html"
	if from_disk and D.exists():
		with X(D,'r')as G:C=G.read()
		J=B.search(P,C)
		if not J:return C
	K=g.firefox();C=Q(E);z(C)
	if D:
		with X(D,'w',encoding='ISO-8859-15')as G:G.write(C)
	return C
def A5(html):
	H=i(html,q);D=H.find(e,class_=f);I=D.find_all('font',class_='visupetit');E=D.find_all('a',class_='listegras')
	for(F,J)in n(I):K='^^([a-zA-Z]*) ?(\\d)?';C=B.search(K,E[F].text);L=C.group(2)if C and C.lastindex>=2 else A;C=C.group(1)if C else A;G.append([J.text,C+' '+L,E[F].text.replace(';',',').strip()])
def A6(html,num_module):
	H=num_module;M=i(html,q);N=M.find_all(e,class_=f)[1];P=N.find_all('td',class_=lambda c:c!=f and(c=='listeelem1'or c=='listeelem2'));E=0;F=A;J=A
	for C in P:
		Q=C.find(e)
		if Q:C=C.find_all('td')[0]
		if C.text and'<a href='not in C:
			K=B.findall('^\\d{4}',C.text)
			if K:L=B.findall('(\\d) ?H$',C.text);G='^\\d{4} *-? *(.+?)( *-? *\\d *H)?$';R=B.search(G,C.text);F=K[0];J=L[0]if L else A;O.append([H,F,R.group(1).replace(';',',').strip(),J]);E=0
			elif p in C.text:E=C.text[-2:]
			else:
				G='^(\\d{2}\\/\\d{2}\\/\\d{4})\xa0(\\d{2}:\\d{2})\xa0\xa0(\\d{2}\\/\\d{2}\\/\\d{4})\xa0(\\d{2}:\\d{2})$';D=B.search(G,C.text)
				if D:S=D.group(1)if D else A;T=D.group(2)if D and D.lastindex>=2 else A;U=D.group(4)if D and D.lastindex>=4 else A;I.append([H,F,E,S,T,U])
if __name__=='__main__':
	H=M('cache')
	if not H.exists():H.mkdir(parents=J)
	H=M('output')
	if not H.exists():H.mkdir(parents=J)
	k=y();D();A7=j(N[k],from_disk=d);A5(A7)
	for(A8,A9)in n(G[1:]):v(A8+1,W(G[1:]),title='Lecture des modules de formation (MAJ dates)        ');l=A9[0];AA=w[k]+l;AB=j(AA,from_disk=d);A6(AB,l)
	D();D();A4('output/synthese_planning.xlsx');D();D("Traitement terminé, le fichier de synthèse se trouve dans le dossier 'output'");m('Appuyez sur Entrée pour fermer le programme...')
