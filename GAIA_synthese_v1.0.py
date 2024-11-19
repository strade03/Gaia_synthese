AB='html.parser'
A5='candidat_module_choix_etat'
A4='validation'
A3='pageActionSuivante'
A2='pageAction'
A1='cadrenormal'
A0='table'
z='ISO-8859-15'
y='r'
x='E'
w='Groupe'
v='titre module'
u=range
t=float
i='Dispositifs'
h='D'
g='C'
f='B'
e='A'
d='w'
c='N° module'
b='N° dispo.'
a=int
T=enumerate
S=input
K=False
J='Inscriptions'
I='Modules'
H=open
G=len
F='Dates'
D=True
C=print
A=''
import re as B
from pathlib import Path as L
import pandas as E
from openpyxl import load_workbook as AC
from openpyxl.styles import Border,Side,Alignment as AD
import time
from datetime import datetime as AE
import sys,math,browser_cookie3 as j,requests as U
from bs4 import BeautifulSoup as A6
def k(step,total_steps,bar_width=60,title=A,print_perc=D):
	F=title;import sys;H=['█','▏','▎','▍','▌','▋','▊','█'];C=100*t(step)/t(total_steps);I=bar_width*8;E=a(round(C/100*I));K=E/8;J=E%8;B=D=A;D+=H[0]*a(K)
	if J>0:D+=H[J]
	D+='▒'*a(I/8-t(E)/8.)
	if G(F)>0:B=F+': '
	B+='\x1b[0;32m';B+=D;B+='\x1b[0m'
	if print_perc:
		if C>1e2:C=1e2
		B+=' {:6.2f}'.format(C)+' %'
	sys.stdout.write('\r'+B);sys.stdout.flush()
l=['https://gaia.phm.education.gouv.fr/gaia/gacmfgest/animPedagogiques/liste_dispoAnim.jsp','https://gaia.in.phm.education.gouv.fr/gaia/gacmfgest/animPedagogiques/liste_dispoAnim.jsp']
AF=['https://gaia.phm.education.gouv.fr/gaia/gacmfgest/dispo/arbre_dispositif.jsp?cCode=','https://gaia.in.phm.education.gouv.fr/gaia/gacmfgest/dispo/arbre_dispositif.jsp?cCode=']
A7=['https://gaia.phm.education.gouv.fr/gaia/gacmfgest/centrale','https://gaia.in.phm.education.gouv.fr/gaia/gacmfgest/centrale']
m={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:131.0) Gecko/20100101 Firefox/131.0'}
M=[[b,'Circo','titre dispo.']]
N=[[b,c,v,"nombre d'heures",'nb inscrits']]
V=[[b,c,w,'date','Heure début','Heure fin']]
O=[[c,w,'Nom']]
n=[]
P='(E R R E U R|authentification)'
def Ad(namefile,data):
	with H(namefile,d,newline='\r\n',encoding='utf-8')as A:
		for B in data:A.write(f"{B}\n")
def o(url):
	C=j.firefox()
	try:A=U.get(url,headers=m,cookies=C);A.raise_for_status();B=A.text
	except U.exceptions.RequestException as D:B='exception ERREUR authentification'
	return B
def AG():
	A=o(l[0])
	if B.search(P,A):
		A=o(l[1])
		if B.search('exception',A):return 0
		else:return 1
	return 0
def p(page):
	A=B.search(P,page)
	if A:C('\nERREUR, ouvrez une session GAIA sous Firefox, vous êtes actuellement non connecté');C("Si l'erreur persiste cliquer dans Stagiaires=>Personnes ou naviguer sur le site GAIA");S('\nAppuyez sur Entrée pour fermer le programme...');sys.exit()
def AH(wb):A=wb[F];A.column_dimensions[e].width=13;A.column_dimensions[f].width=14;A.column_dimensions[g].width=12;A.column_dimensions[h].width=11;A.column_dimensions[x].width=12;A.column_dimensions['F'].width=12;A.column_dimensions['G'].width=60;A.column_dimensions['H'].width=14
def AI(wb):A=wb[i];A.column_dimensions[e].width=13;A.column_dimensions[f].width=14;A.column_dimensions[g].width=80
def AJ(wb):A=wb[I];A.column_dimensions[e].width=13;A.column_dimensions[f].width=10;A.column_dimensions[g].width=70;A.column_dimensions[h].width=16;A.column_dimensions[x].width=10
def AK(wb):A=wb[J];A.column_dimensions[e].width=10;A.column_dimensions[f].width=10;A.column_dimensions[g].width=70;A.column_dimensions[h].width=60;A.column_dimensions[x].width=14
def W(ws,first_row,first_col,last_col):
	for A in ws.iter_rows(min_row=first_row,max_row=ws.max_row,min_col=first_col,max_col=last_col):
		for B in A:B.alignment=AD(horizontal='center')
def X(ws,first_row,first_col):
	A=Side(border_style='thin')
	for B in ws.iter_rows(min_row=first_row,max_row=ws.max_row,min_col=first_col,max_col=ws.max_column):
		for C in B:C.border=Border(left=A,right=A,top=A,bottom=A)
def Y(ws,ref):A=ws.auto_filter;A.ref=ref
def AL(ws,col):
	for A in ws[col]:
		if A.row>1:A.number_format='DD/MM/YYYY'
def AM(df,num_module):
	B=df.loc[df[c]==num_module,b]
	if not B.empty:return B.iloc[0]
	return A
def AN(namefile):
	R='A:E';Q='circo';L=namefile;P=V[1:];P.sort(key=lambda x:AE.strptime(x[3],'%d/%m/%Y'));B={i:E.DataFrame(M[1:],columns=M[0]),I:E.DataFrame(N[1:],columns=N[0]),F:E.DataFrame(P,columns=V[0]),J:E.DataFrame(O[1:],columns=O[0])};C=[];D=[]
	for H in u(1,G(V)):C.append(f"=VLOOKUP(B{H+1},Modules!B:D,2,FALSE)");D.append(f"=VLOOKUP(A{H+1},Dispositifs!A:C,2,FALSE)")
	B[F][v]=E.Series(C);B[F][Q]=E.Series(D);C=[];D=[]
	for H in u(1,G(O)):C.append(f"=VLOOKUP(A{H+1},Modules!B:D,2,FALSE)");D.append(f'=VLOOKUP("{AM(B[I],O[H][0])}",Dispositifs!A:C,2,FALSE)')
	B[J][v]=E.Series(C);B[J][Q]=E.Series(D)
	with E.ExcelWriter(L,engine='openpyxl')as S:
		for(T,U)in B.items():U.to_excel(S,sheet_name=T,index=K)
	A=AC(L);AH(A);W(A[F],1,2,6);X(A[F],1,1);AL(A[F],h);Y(A[F],'A:H');AI(A);X(A[i],1,1);Y(A[i],'A:C');AJ(A);W(A[I],1,2,2);W(A[I],1,4,5);X(A[I],1,1);Y(A[I],R);AK(A);W(A[J],1,2,2);X(A[J],1,1);Y(A[J],R);A.save(L)
def A8(url,from_disk=D):
	E=url;D=A;I=B.search('/([^/]+)\\.jsp\\S*',E);J=B.findall('=([^/]+)',E)
	if I:
		F='cache/'+I.group(1)
		if J:F=F+'_'+J[0]
		D=L(__file__).parent/f"{F}.html"
	if from_disk and D.exists():
		with H(D,y)as G:C=G.read()
		K=B.search(P,C)
		if not K:return C
	C=o(E);p(C)
	if D:
		with H(D,d,encoding=z)as G:G.write(C)
	return C
def AO(html):
	G=A6(html,AB);D=G.find(A0,class_=A1);H=D.find_all('font',class_='visupetit');E=D.find_all('a',class_='listegras')
	for(F,I)in T(H):J='^^([a-zA-Z]*) ?(\\d)?';C=B.search(J,E[F].text);K=C.group(2)if C and C.lastindex>=2 else A;C=C.group(1)if C else A;M.append([I.text,C+' '+K,E[F].text.replace(';',',').strip()])
def AP(html,num_module):
	H=num_module;L=A6(html,AB);M=L.find_all(A0,class_=A1)[1];O=M.find_all('td',class_=lambda c:c!=A1 and(c=='listeelem1'or c=='listeelem2'));E=0;F=A;I=A
	for C in O:
		P=C.find(A0)
		if P:C=C.find_all('td')[0]
		if C.text and'<a href='not in C:
			J=B.findall('^\\d{4}',C.text)
			if J:K=B.findall('(\\d) ?H$',C.text);G='^\\d{4} *-? *(.+?)( *-? *\\d *H)?$';Q=B.search(G,C.text);F=J[0];I=K[0]if K else A;N.append([H,F,Q.group(1).replace(';',',').strip(),I,0]);E=0
			elif w in C.text:E=C.text[-2:]
			else:
				G='^(\\d{2}\\/\\d{2}\\/\\d{4})\xa0(\\d{2}:\\d{2})\xa0\xa0(\\d{2}\\/\\d{2}\\/\\d{4})\xa0(\\d{2}:\\d{2})$';D=B.search(G,C.text)
				if D:R=D.group(1)if D else A;S=D.group(2)if D and D.lastindex>=2 else A;T=D.group(4)if D and D.lastindex>=4 else A;V.append([H,F,E,R,S,T])
def AQ(url,code_dispo,code_module,from_disk=D):
	M=code_dispo;G=code_module;E=L(__file__).parent/f"cache/coDisp_{G}.html";N=D
	if from_disk and E.exists():
		with H(E,y)as I:C=I.read()
		R=B.search(P,C)
		if not R:N=K
	O='20'+M[0:2];id=M[0:6];S={'etatCandidatAccueil':A4,'modeCandidatAccueil':'choixmodule','pageActionDispoUnique':A,'pageActionPersonneUnique':A,'pageActionPersonneUniqueSansCandidature':A,'pageActionDispoUniqueSansCandidature':A,A2:'candidat_accueil',A3:A5,'sDispoPrefix':id,'modesansactif':'Y','champ_id':id,'champ_annee_dispositif':O,'champ_libl_dispositif':A,'champ_co_modu':f"{G}",'champ_annee_personne':O,'champ_nom_personne':A,'champ_prenom_personne':A,'champ_code_etab':A,'filtrePersActif':'on'}
	if N:
		T=j.firefox();V=U.post(url,headers=m,cookies=T,data=S);C=V.text;p(C)
		if E:
			with H(E,d,encoding=z)as I:I.write(C)
	J='coDisp" value="(\\d*)"';F=B.search(J,C);F=F.group(1)if F else'nop';J='TOUTES <\\/b> \\((\\d*)';Q=B.search(J,C);W=a(Q.group(1))if Q else 0;return[G,F,W]
def AR(url,num_module,num_coDisp,nb_page,from_disk=D):
	Y='champ_etat';S=num_coDisp;R='candidat_module_liste_candidature';M='T';F=num_module;V=L(__file__).parent/f"cache/inscrits_{F}"
	for N in u(nb_page):
		I=V.with_name(f"{V.name}-{N}.html");W=D
		if from_disk and I.exists():
			with H(I,y)as Q:C=Q.read()
			Z=B.search(P,C)
			if not Z:W=K
		if N==0:X={'etatCandidatModuleChoixEtat':A4,'modeCandidatModuleChoixEtat':'selection',A2:A5,A3:R,'etat':M,'coModu':F,'coDisp':S,Y:M}
		else:X={'etatCandidatModuleListeCandidature':A4,'modeCandidatModuleListeCandidature':'avanceposition',A2:R,A3:R,'pageActionPrecedente':A5,'champ_noMatr':A,'champ_noMatrVisuPersonne':A,'champ_coDisp':S,'champ_coModu':F,'etatChoisi':M,Y:M,'champ_lettre':A,'champ_reculeposition':A,'champ_avanceposition':f"{15*N}",'champ_positiondepart':'0'}
		if W:
			a=j.firefox();time.sleep(.5);b=U.post(url,headers=m,cookies=a,data=X);C=b.text;p(C)
			if I:
				with H(I,d,encoding=z)as Q:Q.write(C)
		E='visualisation_personne\\(\'\\d*\'\\);\\">\\s*([a-zA-Z .-]*)<\\/a>';c=B.findall(E,C);E="visualisation_personne\\('(\\d*)'\\)";e=B.findall(E,C)
		for(f,g)in T(c):E=f'name=\\"candidature{e[f]}GroupeChoisi\\"\\s*value=\\"(\\d*)';J=B.findall(E,C);J=J[0]if G(J)>0 else A;O.append([F,J,g])
if __name__=='__main__':
	Q=L('cache')
	if not Q.exists():Q.mkdir(parents=D)
	Q=L('output')
	if not Q.exists():Q.mkdir(parents=D)
	Z=AG();C("Un système de cache permet d'accèlerer le traitement mais ne mettra à jour que les nouvelles informations");C("Si vous demandez la mise à jour, les informations seront récupérées sur GAIA sinon à partir du cache s'il existe.");C('Que souhaitez-vous faire ?');AS=S('Mettre à jour les dates des formations (O/N) :');q=S("Mettre à jour le nombre d'inscriptions (O/N) : ");AT=S('Mettre à jour la listes des enseignants inscrits (O/N) :');AU=K if AS.lower()=='o'else D;AV=K if q.lower()=='o'else D;AW=K if AT.lower()=='o'else D;C();AX=A8(l[Z],from_disk=K);AO(AX)
	for(R,q)in T(M[1:]):k(R+1,G(M[1:]),title='Lecture des modules de formation (MAJ dates)        ');A9=q[0];AY=AF[Z]+A9;AZ=A8(AY,from_disk=AU);AP(AZ,A9)
	C()
	for(R,r)in T(N[1:]):
		k(R+1,G(N)-1,title="Lecture des dispositifs (MAJ nb d'inscrits)         ");s=AQ(A7[Z],r[0],r[1],from_disk=AV)
		if'nop'not in s:r[4]=s[2];n.append(s)
	C()
	for(R,(Aa,Ab,AA))in T(n):
		k(R+1,G(n)-1,title='Lecture des inscriptions (liste des inscriptions)   ')
		if AA>0:Ac=math.ceil(AA/15);AR(A7[Z],Aa,Ab,Ac,from_disk=AW)
	C();AN('output/synthese.xlsx');C();C("Traitement terminé, le fichier de synthèse se trouve dans le dossier 'output'");S('Appuyez sur Entrée pour fermer le programme...')
